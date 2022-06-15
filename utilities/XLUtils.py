import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

def dataValidation(file,sheetName,checkCell):
    row = getRowCount(file,sheetName)
    col = getColumnCount(file,sheetName)
    for r in range(1, row + 1):
        for c in range(1,col + 1):
            exp = readData(file,sheetName, r, c)
            if checkCell == exp:
                status = "True"
                return status

def readData(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum,columnno).value

def writeData(file,sheetName,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum,columnno).value=data
    workbook.save(file)

def fillGreenColor(file,sheetname,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = patternFill(start_color='60b212',end_color='60b212',fill_type='solid')
    sheet.cell(rownum,columnno).fill=greenFill
    workbook.save(file)

def fillRedcolor(file,sheetName,rownum,columnno):
    redFill=patternFill(start_color='ff0000',end_color='ff0000',fill_type='solid')
    workbook.save(file)


